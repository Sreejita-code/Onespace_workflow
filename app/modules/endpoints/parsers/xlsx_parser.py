import io
import openpyxl
from . import _try_json

def parse_xlsx(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    # Get headers from the first row
    headers = [str(c.value).lower().replace(" ", "_") if c.value else f"col_{i}" 
               for i, c in enumerate(next(ws.iter_rows()))]
    
    endpoints = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue # Skip entirely empty rows
        
        record = dict(zip(headers, row))
        endpoints.append({
            "method":       str(record.get("method", "GET")).upper(),
            "url":          str(record.get("url", "")),
            "description":  str(record.get("description", "")),
            "request_body": _try_json(record.get("request_body")),
            "response_body":_try_json(record.get("response_body")),
        })
    return endpoints